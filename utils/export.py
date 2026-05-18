import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import make_response

def export_csv(headers, rows, filename="export.csv"):
    si = io.StringIO()
    writer = csv.writer(si)
    writer.writerow(headers)
    writer.writerows(rows)
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "text/csv"
    return output

def export_xlsx(headers, rows, sheet_name="Data", filename="export.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    output = make_response(buf.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return output
